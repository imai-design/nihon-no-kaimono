"""経済産業省「企業活動基本調査」2025年確報（2024年度実績）から
サイト表示用の kigyo.json を生成するスクリプト。

入力（LibreOfficeで.xlsから変換済みの.xlsx。ダウンロード不要）:
  data/kigyo/xlsx_converted/000040464647.xlsx  第1表 産業別 総括表
  data/kigyo/xlsx_converted/000040464650.xlsx  第4表 産業別×資本金規模別
  data/kigyo/xlsx_converted/000040464656.xlsx  第10表 産業別 研究開発費

出力:
  site/kigyo.json
"""

import json
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_DIR = BASE_DIR / "data" / "kigyo" / "xlsx_converted"
OUTPUT_PATH = BASE_DIR / "site" / "kigyo.json"

TABLE1_PATH = XLSX_DIR / "000040464647.xlsx"
TABLE1_SHEET = "2025_kaku_01-01"
TABLE4_PATH = XLSX_DIR / "000040464650.xlsx"
TABLE4_SHEET = "2025_kaku_01-04"
TABLE10_PATH = XLSX_DIR / "000040464656.xlsx"
TABLE10_SHEET = "2025_kaku_01-10"

# 第1表 データ行の開始行（1〜5行目はタイトル・見出し）
TABLE1_DATA_START_ROW = 6
# 第1表 列インデックス（1始まり。実データのヘッダー文字列で検証済み）
TABLE1_COL = {
    "industry_code": 1,
    "industry_name": 2,
    "fiscal_year": 4,
    "companies": 5,
    "employees": 7,
    "sales": 16,
    "operating_profit": 19,
    "ordinary_profit": 20,
    "value_added": 22,
    "subsidiaries_total": 23,
    "subsidiaries_domestic": 24,
    "subsidiaries_overseas": 25,
}

# 第4表 データ行の開始行
TABLE4_DATA_START_ROW = 6
TABLE4_COL = {
    "industry_code": 1,
    "industry_name": 2,
    "capital_band_code": 3,
    "capital_band_name": 4,
    "fiscal_year": 6,
    "companies": 7,
    "sales": 8,
    "ordinary_profit": 9,
    "value_added": 11,
    "operating_profit": 24,
}
# 資本金規模コード "00"（計）は全産業の総合計と同じなので by_capital には含めない
TABLE4_CAPITAL_TOTAL_BAND_CODE = "00"

# 第10表 データ行の開始行（1〜8行目はタイトル・見出し）
TABLE10_DATA_START_ROW = 9
TABLE10_COL = {
    "industry_code": 1,
    "industry_name": 2,
    "fiscal_year": 4,
    "rd_ratio": 12,
}

# 秘匿・非該当を表す記号
SECRECY_MARKERS = {"***", "X", "x", "-"}

MONEY_UNIT = "百万円"
PEOPLE_UNIT = "人"
COMPANIES_UNIT = "社"
SCOPE_NOTE = "この調査の対象は従業者50人以上かつ資本金3,000万円以上の会社です。"

SOURCES = [
    {
        "name": "経済産業省企業活動基本調査 2025年確報（2024年度実績）第1表",
        "statInfId": "000040464647",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040464647&fileKind=0",
    },
    {
        "name": "同 第10表（研究開発費）",
        "statInfId": "000040464656",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040464656&fileKind=0",
    },
    {
        "name": "同 第4表（資本金規模別）",
        "statInfId": "000040464650",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040464650&fileKind=0",
    },
]

TOTAL_CODES = {"総合計", "000"}
MINOR_CODE_PATTERN = re.compile(r"^\d+$")


def classify_level(industry_code):
    """産業分類コードから level を機械的に判定する。"""
    code = str(industry_code)
    if code in TOTAL_CODES:
        return "total"
    if MINOR_CODE_PATTERN.fullmatch(code):
        return "minor"
    return "major"


def clean_number(value):
    """秘匿記号・空欄を None に、それ以外は数値として返す。"""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in SECRECY_MARKERS or stripped == "":
            return None
        # 万一文字列で数値が入っていた場合に備えて変換を試みる
        try:
            return float(stripped) if "." in stripped else int(stripped)
        except ValueError:
            return None
    return value


def find_latest_fiscal_year(ws, year_col, data_start_row):
    """年度列から最新の年度文字列（例: "2024年度"）を機械的に検出する。"""
    years = set()
    for row in ws.iter_rows(min_row=data_start_row, values_only=False):
        cell = row[year_col - 1]
        value = cell.value
        if isinstance(value, str) and value.endswith("年度"):
            years.add(value)
    if not years:
        raise ValueError("年度列から年度を検出できませんでした")

    def year_number(year_str):
        return int(year_str.replace("年度", ""))

    return max(years, key=year_number)


def load_table1_industries():
    wb = openpyxl.load_workbook(TABLE1_PATH, data_only=True)
    ws = wb[TABLE1_SHEET]
    col = TABLE1_COL
    latest_year = find_latest_fiscal_year(ws, col["fiscal_year"], TABLE1_DATA_START_ROW)

    industries = []
    for r in range(TABLE1_DATA_START_ROW, ws.max_row + 1):
        fiscal_year = ws.cell(row=r, column=col["fiscal_year"]).value
        if fiscal_year != latest_year:
            continue
        industry_code = ws.cell(row=r, column=col["industry_code"]).value
        if industry_code is None:
            continue
        industry_name = ws.cell(row=r, column=col["industry_name"]).value

        industries.append(
            {
                "code": industry_code,
                "name": industry_name,
                "level": classify_level(industry_code),
                "companies": clean_number(ws.cell(row=r, column=col["companies"]).value),
                "employees": clean_number(ws.cell(row=r, column=col["employees"]).value),
                "sales": clean_number(ws.cell(row=r, column=col["sales"]).value),
                "operating_profit": clean_number(
                    ws.cell(row=r, column=col["operating_profit"]).value
                ),
                "ordinary_profit": clean_number(
                    ws.cell(row=r, column=col["ordinary_profit"]).value
                ),
                "value_added": clean_number(ws.cell(row=r, column=col["value_added"]).value),
                "subsidiaries_total": clean_number(
                    ws.cell(row=r, column=col["subsidiaries_total"]).value
                ),
                "subsidiaries_domestic": clean_number(
                    ws.cell(row=r, column=col["subsidiaries_domestic"]).value
                ),
                "subsidiaries_overseas": clean_number(
                    ws.cell(row=r, column=col["subsidiaries_overseas"]).value
                ),
                "rd_ratio": None,  # 後で第10表から業種名で突き合わせて埋める
            }
        )
    return industries, latest_year


def load_table10_rd_ratio_by_name(latest_year):
    """第10表から 業種名 -> 売上高研究開発費比率 の辞書を作る。"""
    wb = openpyxl.load_workbook(TABLE10_PATH, data_only=True)
    ws = wb[TABLE10_SHEET]
    col = TABLE10_COL

    rd_ratio_by_name = {}
    for r in range(TABLE10_DATA_START_ROW, ws.max_row + 1):
        fiscal_year = ws.cell(row=r, column=col["fiscal_year"]).value
        if fiscal_year != latest_year:
            continue
        industry_name = ws.cell(row=r, column=col["industry_name"]).value
        if industry_name is None:
            continue
        rd_ratio = clean_number(ws.cell(row=r, column=col["rd_ratio"]).value)
        # 同名で複数行ある場合（データ上の重複行）は最初に見つかった値を採用する
        rd_ratio_by_name.setdefault(industry_name, rd_ratio)
    return rd_ratio_by_name


def load_table4_by_capital(latest_year, major_industry_names):
    """第4表から 大分類業種名 -> 資本金規模別リスト の辞書を作る。"""
    wb = openpyxl.load_workbook(TABLE4_PATH, data_only=True)
    ws = wb[TABLE4_SHEET]
    col = TABLE4_COL

    by_capital = {}
    for r in range(TABLE4_DATA_START_ROW, ws.max_row + 1):
        fiscal_year = ws.cell(row=r, column=col["fiscal_year"]).value
        if fiscal_year != latest_year:
            continue
        industry_name = ws.cell(row=r, column=col["industry_name"]).value
        if industry_name not in major_industry_names:
            continue
        capital_band_code = ws.cell(row=r, column=col["capital_band_code"]).value
        if capital_band_code == TABLE4_CAPITAL_TOTAL_BAND_CODE:
            continue

        capital_band_name = ws.cell(row=r, column=col["capital_band_name"]).value
        entry = {
            "band": capital_band_name,
            "companies": clean_number(ws.cell(row=r, column=col["companies"]).value),
            "sales": clean_number(ws.cell(row=r, column=col["sales"]).value),
            "operating_profit": clean_number(
                ws.cell(row=r, column=col["operating_profit"]).value
            ),
            "ordinary_profit": clean_number(
                ws.cell(row=r, column=col["ordinary_profit"]).value
            ),
            "value_added": clean_number(ws.cell(row=r, column=col["value_added"]).value),
        }
        by_capital.setdefault(industry_name, []).append(entry)
    return by_capital


def build():
    industries, latest_year = load_table1_industries()

    rd_ratio_by_name = load_table10_rd_ratio_by_name(latest_year)
    for industry in industries:
        industry["rd_ratio"] = rd_ratio_by_name.get(industry["name"])

    major_industry_names = {ind["name"] for ind in industries if ind["level"] == "major"}
    by_capital = load_table4_by_capital(latest_year, major_industry_names)

    output = {
        "meta": {
            "year": latest_year,
            "generated_at": "2026-08-08",
            "unit": {"money": MONEY_UNIT, "people": PEOPLE_UNIT, "companies": COMPANIES_UNIT},
            "sources": SOURCES,
            "scope_note": SCOPE_NOTE,
        },
        "industries": industries,
        "by_capital": by_capital,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    return output


if __name__ == "__main__":
    result = build()
    print(f"wrote {OUTPUT_PATH}")
    print(f"industries: {len(result['industries'])}")
    print(f"by_capital industries: {len(result['by_capital'])}")

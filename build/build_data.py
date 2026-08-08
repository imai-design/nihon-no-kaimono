#!/usr/bin/env python3
"""経済産業省「商業動態統計調査」の公表Excelを、サイト表示用のJSONに変換する。

入力: data/raw/<statInfId>.xls  (拡張子は .xls だが実体は OOXML/xlsx)
出力: site/data.json

方針:
- ヘッダー行・列位置はハードコードせず、シート内の「時間軸コード」ラベルと
  各系列の見出し文字列（Total / Sales of goods / Retail 等）を検索して特定する。
  シートごとにレイアウトの微妙な差異（先頭に空列があるかどうか等）があるため。
- 欠測（"***" やセル空欄）は None として出力する。推計で埋めない。
- 前年同月比は、公表シートの値（100=前年同月と同水準、107.6=+7.6%）から
  100 を引いて % 変化量に変換する。
"""
import io
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "data" / "raw"
OUT_PATH = BASE_DIR / "site" / "data.json"

UNIT = "百万円"
MISSING_MARKERS = {"", "***"}

# 6業態すべてが揃う開始月
MONTHS_START_YEAR = 2014
MONTHS_START_MONTH = 1

HEADER_SEARCH_MAX_ROW = 20
HEADER_SEARCH_MAX_COL = 5

BILLION_TO_MILLION = 1000  # 10億円 -> 百万円

SOURCES = [
    {
        "name": "商業動態統計 業種別商業販売額（小売業計）",
        "statInfId": "000031387992",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000031387992&fileKind=0",
    },
    {
        "name": "商業動態統計 百貨店・スーパー商品別販売額",
        "statInfId": "000031387994",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000031387994&fileKind=0",
    },
    {
        "name": "商業動態統計 コンビニエンスストア販売額等",
        "statInfId": "000031387995",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000031387995&fileKind=0",
    },
    {
        "name": "商業動態統計 家電大型専門店販売",
        "statInfId": "000031387996",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000031387996&fileKind=0",
    },
    {
        "name": "商業動態統計 ドラッグストア販売",
        "statInfId": "000031387997",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000031387997&fileKind=0",
    },
    {
        "name": "商業動態統計 ホームセンター販売",
        "statInfId": "000031387998",
        "url": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000031387998&fileKind=0",
    },
]

# 各系列の定義。value/yoy とも「シート名」と「見出し文字列(label)」で列を特定する。
SERIES_CONFIG = [
    {
        "key": "department",
        "name": "百貨店",
        "short": "百貨店",
        "note": "全国百貨店の販売額",
        "value": {"stat": "000031387994", "sheet": "百貨店　販売額　月次", "label": "Total"},
        "yoy": {"stat": "000031387994", "sheet": "百貨店　伸び率　月次", "label": "Total"},
    },
    {
        "key": "supermarket",
        "name": "スーパー",
        "short": "スーパー",
        "note": "総合スーパー・食品スーパーなど",
        "value": {"stat": "000031387994", "sheet": "スーパー　販売額　月次", "label": "Total"},
        "yoy": {"stat": "000031387994", "sheet": "スーパー　伸び率　月次", "label": "Total"},
    },
    {
        "key": "cvs",
        "name": "コンビニエンスストア",
        "short": "コンビニ",
        "note": "主要コンビニエンスストアチェーンの販売額",
        "value": {"stat": "000031387995", "sheet": "販売額(value)月次(Monthly)", "label": "Total"},
        "yoy": {"stat": "000031387995", "sheet": "前年同月比(change) (Monthly)", "label": "Total"},
    },
    {
        "key": "kaden",
        "name": "家電大型専門店",
        "short": "家電量販店",
        "note": "家電大型専門店（量販店）の販売額",
        "value": {"stat": "000031387996", "sheet": "販売額(value)月次(Monthly)", "label": "Sales of goods"},
        "yoy": {"stat": "000031387996", "sheet": "前年同月比(change) (Monthly)", "label": "Sales of goods"},
    },
    {
        "key": "drug",
        "name": "ドラッグストア",
        "short": "ドラッグストア",
        "note": "ドラッグストアの販売額",
        "value": {"stat": "000031387997", "sheet": "販売額(value)月次(Monthly)", "label": "Sales of goods"},
        "yoy": {"stat": "000031387997", "sheet": "前年同月比(change) (Monthly)", "label": "Sales of goods"},
    },
    {
        "key": "homecenter",
        "name": "ホームセンター",
        "short": "ホームセンター",
        "note": "ホームセンターの販売額",
        "value": {"stat": "000031387998", "sheet": "販売額(value)月次(Monthly)", "label": "Sales of goods"},
        "yoy": {"stat": "000031387998", "sheet": "前年同月比(change) (Monthly)", "label": "Sales of goods"},
    },
]

RETAIL_TOTAL_CONFIG = {
    "value": {"stat": "000031387992", "sheet": "販売額（value）(月次M)", "label": "Retail", "multiplier": BILLION_TO_MILLION},
    "yoy": {"stat": "000031387992", "sheet": "前年比（change）(月次M)", "label": "Retail"},
}

LONGRUN_CONFIG = {
    "department": {"stat": "000031387994", "sheet": "百貨店　販売額　年", "label": "Total"},
    "supermarket": {"stat": "000031387994", "sheet": "スーパー　販売額　年", "label": "Total"},
}

MONTH_PATTERN = re.compile(r"(\d{4})年\s*(\d{1,2})月")
YEAR_PATTERN = re.compile(r"(\d{4})年")

_workbook_cache = {}


def load_workbook(stat_inf_id):
    """statInfId の xlsx(拡張子.xls)を読み込む。同一プロセス内でキャッシュする。"""
    if stat_inf_id in _workbook_cache:
        return _workbook_cache[stat_inf_id]
    path = RAW_DIR / f"{stat_inf_id}.xls"
    if not path.exists():
        raise FileNotFoundError(f"raw file not found: {path}")
    with open(path, "rb") as fh:
        data = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    _workbook_cache[stat_inf_id] = wb
    return wb


def find_header(ws):
    """「時間軸コード」ラベルのセルを探し、(header_row, time_code_col) を返す。

    シートによって先頭に空列が挿入されているものがあるため、列も検索する。
    """
    for r in range(1, HEADER_SEARCH_MAX_ROW + 1):
        for c in range(1, HEADER_SEARCH_MAX_COL + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == "時間軸コード":
                return r, c
    raise ValueError(f"header row (時間軸コード) not found in sheet '{ws.title}'")


def find_col_by_label(ws, header_row, label):
    """ヘッダー行から label(大文字小文字無視)に完全一致する列番号(1始まり)を返す。"""
    target = label.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip().lower() == target:
            return c
    raise ValueError(f"column label '{label}' not found in header row {header_row} of sheet '{ws.title}'")


def to_number(raw):
    """セルの値を数値(int/float)またはNoneに変換する。欠測(***, 空欄)はNone。"""
    if raw is None:
        return None
    if isinstance(raw, str):
        s = raw.strip()
        if s in MISSING_MARKERS:
            return None
        try:
            return float(s) if "." in s else int(s)
        except ValueError:
            return None
    return raw


def extract_monthly_series(stat_inf_id, sheet_name, label, multiplier=1):
    """月次シートから {"YYYY-MM": 値|None} の辞書を返す（時系列順）。"""
    wb = load_workbook(stat_inf_id)
    ws = wb[sheet_name]
    header_row, time_code_col = find_header(ws)
    year_month_col = time_code_col + 1
    value_col = find_col_by_label(ws, header_row, label)

    result = {}
    for r in range(header_row + 1, ws.max_row + 1):
        ym_cell = ws.cell(row=r, column=year_month_col).value
        if not isinstance(ym_cell, str):
            continue
        m = MONTH_PATTERN.search(ym_cell)
        if not m:
            continue
        year, month = int(m.group(1)), int(m.group(2))
        key = f"{year:04d}-{month:02d}"
        val = to_number(ws.cell(row=r, column=value_col).value)
        if val is not None and multiplier != 1:
            val = val * multiplier
        result[key] = val
    return result


def extract_annual_series(stat_inf_id, sheet_name, label):
    """年次シートから {year(int): 値|None} の辞書を返す。"""
    wb = load_workbook(stat_inf_id)
    ws = wb[sheet_name]
    header_row, time_code_col = find_header(ws)
    year_col = time_code_col + 1
    value_col = find_col_by_label(ws, header_row, label)

    result = {}
    for r in range(header_row + 1, ws.max_row + 1):
        y_cell = ws.cell(row=r, column=year_col).value
        if not isinstance(y_cell, str):
            continue
        m = YEAR_PATTERN.search(y_cell)
        if not m:
            continue
        year = int(m.group(1))
        val = to_number(ws.cell(row=r, column=value_col).value)
        result[year] = val
    return result


def ratio_to_pct_change(ratio):
    """公表の前年比指数(100=横ばい)を % 変化量に変換する。欠測はNoneのまま。"""
    if ratio is None:
        return None
    return round(ratio - 100, 1)


def month_range(start_year, start_month, end_year, end_month):
    months = []
    y, m = start_year, start_month
    while (y, m) <= (end_year, end_month):
        months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def build():
    # --- 各系列の月次販売額・前年同月比を取得 ---
    series_values = {}
    series_yoy_raw = {}
    latest_months = []

    for cfg in SERIES_CONFIG:
        key = cfg["key"]
        v_cfg = cfg["value"]
        values_dict = extract_monthly_series(
            v_cfg["stat"], v_cfg["sheet"], v_cfg["label"], v_cfg.get("multiplier", 1)
        )
        series_values[key] = values_dict
        latest_months.append(max(values_dict.keys()))

        y_cfg = cfg["yoy"]
        yoy_dict = extract_monthly_series(y_cfg["stat"], y_cfg["sheet"], y_cfg["label"])
        series_yoy_raw[key] = yoy_dict

    # --- 業種別商業販売額（小売業計）: 実額か確認済みなので取得 ---
    retail_values = extract_monthly_series(
        RETAIL_TOTAL_CONFIG["value"]["stat"],
        RETAIL_TOTAL_CONFIG["value"]["sheet"],
        RETAIL_TOTAL_CONFIG["value"]["label"],
        RETAIL_TOTAL_CONFIG["value"].get("multiplier", 1),
    )
    latest_months.append(max(retail_values.keys()))

    # --- 6業態＋小売業計すべてに共通する最新月 ---
    latest_month = min(latest_months)

    months = month_range(MONTHS_START_YEAR, MONTHS_START_MONTH, *map(int, latest_month.split("-")))

    values_out = {}
    yoy_out = {}
    latest_value_for_sort = {}

    for cfg in SERIES_CONFIG:
        key = cfg["key"]
        values_out[key] = [series_values[key].get(mth) for mth in months]
        yoy_out[key] = [ratio_to_pct_change(series_yoy_raw[key].get(mth)) for mth in months]
        latest_value_for_sort[key] = series_values[key].get(latest_month)

    retail_total_out = [retail_values.get(mth) for mth in months]

    # --- series を最新月の販売額が大きい順に並べる ---
    ordered_series = sorted(
        SERIES_CONFIG,
        key=lambda c: (latest_value_for_sort[c["key"]] is None, -(latest_value_for_sort[c["key"]] or 0)),
    )
    series_meta = [
        {
            "key": c["key"],
            "name": c["name"],
            "short": c["short"],
            "note": c["note"],
            "yoy_source": "published",
        }
        for c in ordered_series
    ]

    # --- 長期比較（百貨店・スーパー 年次） ---
    dept_annual = extract_annual_series(
        LONGRUN_CONFIG["department"]["stat"], LONGRUN_CONFIG["department"]["sheet"], LONGRUN_CONFIG["department"]["label"]
    )
    sup_annual = extract_annual_series(
        LONGRUN_CONFIG["supermarket"]["stat"], LONGRUN_CONFIG["supermarket"]["sheet"], LONGRUN_CONFIG["supermarket"]["label"]
    )
    years = sorted(set(dept_annual.keys()) & set(sup_annual.keys()))
    longrun = {
        "years": years,
        "department": [dept_annual.get(y) for y in years],
        "supermarket": [sup_annual.get(y) for y in years],
    }

    data = {
        "meta": {
            "latest_month": latest_month,
            "generated_at": date.today().isoformat(),
            "unit": UNIT,
            "sources": SOURCES,
        },
        "series": series_meta,
        "months": months,
        "values": values_out,
        "yoy": yoy_out,
        "retail_total": retail_total_out,
        "longrun": longrun,
    }
    return data


def main():
    data = build()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    print(f"wrote {OUT_PATH}")
    print(f"latest_month={data['meta']['latest_month']}")
    print(f"series order: {[s['key'] for s in data['series']]}")
    print(f"months: {len(data['months'])} ({data['months'][0]} .. {data['months'][-1]})")


if __name__ == "__main__":
    main()
